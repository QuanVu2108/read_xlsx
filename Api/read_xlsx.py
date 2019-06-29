import openpyxl
import os

JSON_MIME_TYPE = 'application/json'

def name_int_chr(col_id):
	if col_id < 27:
		return chr(64 + col_id )
	else:
		return chr(64 + col_id/26) + chr(64 + col_id%26)
	
def name_chr_int(col_id):
	if len(col_id) == 1:
		return ord(col_id) - 64 -1

def get_data(file_id, col_id):
	try :
		dir_path = os.path.dirname(os.path.realpath(__file__))
		xlsx_path = dir_path + '/temp/' + file_id + '.xlsx'
		book = openpyxl.load_workbook(xlsx_path)
		sheet = book.active
		row_min = sheet.min_row
		row_max = sheet.max_row
		col_min = sheet.min_column
		col_max = sheet.max_column
		cells = sheet[name_int_chr(col_min) + str(row_min): name_int_chr(col_max) + str(row_max)]

		data = dict()
		if col_id == 'all':
			for i in range(col_max):
				for j in range(row_max):
					data.update({name_int_chr(i + 1) + str(j + 1) : str(cells[j][i].value)})
					print(type(data))
		elif col_id.isnumeric():
				col_id_int = int(col_id)
				for j in range(row_max):
					data.update({name_int_chr(col_id_int + 1) + str(j + 1) : str(cells[j][int(col_id_int)].value)})
		else :
			for j in range(row_max):
				data.update({col_id + str(j + 1) : str(cells[j][name_chr_int(col_id)].value)})
	except Exception as e:
		data = {}
	return data
		
			
	